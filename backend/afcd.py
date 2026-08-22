"""The Australian Food Composition Database (AFCD), Release 3 — FSANZ's own
published nutrient data — as a local, offline lookup.

This is the difference between "Claude recalls that chicken breast is
roughly 22g of protein" and "here is the FSANZ-published figure for
chicken, breast, lean flesh, raw." The first is a well-informed guess; the
second is the actual reference dataset. `scripts/enrich_pantry.py` tries
this before ever asking a model for a nutrition number.

`scripts/fetch_afcd.py` downloads the two source files FSANZ publishes —
Food details and Nutrient profiles, joined on `Public Food Key` — into
`data/afcd/`, which is **gitignored**: FSANZ states plain copyright with no
licence grant on the download page, so the dataset is not redistributed in
this public repo. Anyone who wants the matching to work runs the fetch
script once, locally.

## Matching

A pantry name is short and informal ("chicken thigh"); an AFCD name is a
comma-separated attribute list ("Chicken, thigh, lean flesh, raw"). Matching
is token-overlap, not a database join — there is no shared key, and no
identifier that survives from a shopping list to a government dataset.

The scoring has one deliberate bias: AFCD carries the same food at several
cooking states (raw / baked / grilled / fried / ...), and a bare pantry
name like "chicken thigh" means *as bought*, not *as cooked*. Candidates
whose name implies a cooking state get a lower score unless the query
itself names one, which is what keeps "chicken thigh" landing on the raw
entry instead of whichever prepared variant happens to score highest on
word overlap alone.

Below `MIN_MATCH_SCORE`, `find_match` returns `None` rather than a weak
guess — the caller's job (see `scripts/enrich_pantry.py`) is to fall back
to an AI estimate for anything this can't confidently place, never to
attach real-looking numbers to a shaky match.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from .nutrition import kcal_from_kj

DATA_DIR = Path(__file__).resolve().parent.parent / "data" / "afcd"
FOOD_DETAILS_FILE = DATA_DIR / "food_details.xlsx"
NUTRIENT_PROFILES_FILE = DATA_DIR / "nutrient_profiles.xlsx"

#: 1-indexed header row in each published sheet (both files share this).
HEADER_ROW = 3

#: Column names in "All solids & liquids per 100 g", exactly as FSANZ
#: publishes them (including the embedded newlines). Chosen to match the
#: values Australian nutrition panels actually display:
#: "Available carbohydrate, without sugar alcohols" is the headline
#: Carbohydrate figure; "Energy with dietary fibre" is the headline Energy
#: figure. Saturated fat is the mass ("g"), not the "%T" fatty-acid-profile
#: column of the same name.
NUTRIENT_COLUMNS = {
    "energy_kj": "Energy with dietary fibre, equated \n(kJ)",
    "protein_g": "Protein \n(g)",
    "fat_g": "Fat, total \n(g)",
    "saturated_fat_g": "Total saturated fatty acids, equated \n(g)",
    "carbs_g": "Available carbohydrate, without sugar alcohols \n(g)",
    "sugars_g": "Total sugars (g)",
    "fibre_g": "Total dietary fibre \n(g)",
    "sodium_mg": "Sodium (Na) \n(mg)",
}


@dataclass(frozen=True)
class AfcdFood:
    key: str
    name: str
    tokens: frozenset[str]
    nutrients: dict[str, float | None]  # energy_kj, protein_g, ... calories_kcal


_TOKEN = re.compile(r"[a-z]+")  # digits dropped: "(~3.5%)" is noise, not signal

#: Cooking states AFCD distinguishes that a bare pantry name never implies
#: — see the module docstring. "dried", "ground", "raw" are left off
#: deliberately: those describe the item *as bought* far more often than
#: a cooking method, so penalising them would push matches the wrong way
#: for spices, flour, and the like.
COOKED_STATES = frozenset(
    {
        "baked", "grilled", "fried", "roasted", "boiled", "steamed",
        "poached", "casseroled", "cooked", "stewed", "braised",
        "microwaved", "canned", "smoked", "barbecued", "stirfried",
        "panfried", "deepfried", "simmered", "toasted",
    }
)

#: Words that describe the *same food in a different state* rather than a
#: different food — cut, preparation, fat content, packaging. An AFCD name
#: carrying one of these that the query didn't ask for costs nothing: a
#: bare "milk" query should still find "Milk, cow, fluid, regular fat"
#: without being penalised for not having said "cow" or "regular".
QUALIFIER_WORDS = (
    frozenset(
        {
            "fresh", "dried", "raw", "peeled", "unpeeled", "trimmed",
            "untrimmed", "whole", "chopped", "sliced", "diced", "grated",
            "crushed", "minced", "ground", "skin", "skinless", "boneless",
            "bonein", "fillet", "flesh", "lean", "fat", "regular",
            "reduced", "low", "high", "light", "full", "standard",
            "natural", "commercial", "homemade", "purchased", "added",
            "without", "with", "salted", "unsalted", "sweetened",
            "unsweetened", "plain", "table", "iodised", "mature", "young",
            "large", "small", "medium", "extra", "organic", "flavour",
            "flavoured", "style", "type", "variety", "dilution",
            "prepared", "from", "dry", "liquid", "cow", "goat",
            "sheep", "no",
        }
    )
    | COOKED_STATES
)

#: Words that mean the candidate is a *different, manufactured* food built
#: from the query ingredient rather than the ingredient itself — this is
#: what "banana" -> "Banana chip" and "carrot" -> "Cake, carrot" were: a
#: short AFCD name with high word-overlap that names the wrong product.
#: Only counts against a candidate when the query didn't ask for it, same
#: as everything else here — "olive oil" still finds "Oil, olive" cleanly
#: because "oil" is then part of the query, not an unexplained extra.
DISTRACTOR_WORDS = frozenset(
    {
        "chip", "chips", "cake", "biscuit", "biscuits", "bread", "juice",
        "oil", "sauce", "jam", "soup", "dip", "spread", "crisp", "crisps",
        "powder",  # milk/coconut/protein powder is a different product,
                   # not a harmless state, from the fluid/whole version
        "syrup", "wine", "beer", "cider", "liqueur", "icecream", "muffin",
        "slice", "pie", "tart", "crumble", "smoothie", "cordial", "drink",
        "milkshake", "chocolate", "lolly", "lollies", "sweet", "sweets",
        "candy", "confectionery", "pudding", "custard", "dessert",
        "snack", "bar", "wafer", "cracker", "crackers", "cereal",
        "beverage", "cocktail", "mocktail", "punch",
    }
)

QUALIFIER_PENALTY = 0.0
DISTRACTOR_PENALTY = 0.35
UNKNOWN_EXTRA_PENALTY = 0.1

MIN_MATCH_SCORE = 0.6
MIN_COVERAGE = 0.8  # nearly every query token has to be found somewhere
COOKED_STATE_PENALTY = 0.6

#: A tie-break nudge, not a real penalty (see LOW_FORM_PENALTY): when a
#: bare name matches several fat-content variants of the same food equally
#: well ("milk" against regular/reduced/skim), prefer the standard one —
#: the reduced version should be found by someone who actually wrote
#: "reduced-fat milk" or "skim milk".
LOW_FORM_WORDS = frozenset({"reduced", "low", "light", "skim", "diet", "fatfree"})
LOW_FORM_PENALTY = 0.02


def _singularise(word: str) -> str:
    if len(word) > 3 and word.endswith("s") and not word.endswith(("ss", "us", "is")):
        return word[:-1]
    return word


def tokenise(text: str) -> frozenset[str]:
    return frozenset(_singularise(t) for t in _TOKEN.findall(text.lower()))


def _score(query_tokens: frozenset[str], food: AfcdFood) -> float:
    """Coverage of the query, minus a penalty for each candidate token the
    query doesn't explain — see QUALIFIER_WORDS / DISTRACTOR_WORDS above.

    Not a symmetric similarity measure on purpose: a longer, more heavily
    qualified AFCD name should never outscore a plainer one just for
    sharing more total words. What matters is whether every extra word in
    the candidate is a harmless qualifier or a sign it's the wrong food.
    """
    if not query_tokens or not food.tokens:
        return 0.0
    overlap = query_tokens & food.tokens
    coverage = len(overlap) / len(query_tokens)
    if coverage < MIN_COVERAGE:
        return 0.0

    penalty = 0.0
    for token in food.tokens - query_tokens:
        if token in QUALIFIER_WORDS:
            penalty += QUALIFIER_PENALTY
        elif token in DISTRACTOR_WORDS:
            penalty += DISTRACTOR_PENALTY
        else:
            penalty += UNKNOWN_EXTRA_PENALTY

    penalty += LOW_FORM_PENALTY * len((food.tokens - query_tokens) & LOW_FORM_WORDS)

    score = max(0.0, coverage - penalty)
    if (food.tokens & COOKED_STATES) and not (query_tokens & COOKED_STATES):
        score *= COOKED_STATE_PENALTY
    return score


def find_match(name: str, foods: list[AfcdFood]) -> tuple[AfcdFood, float] | None:
    """The best AFCD entry for a pantry name, or None below MIN_MATCH_SCORE.

    Ties are broken by shorter AFCD name — the plainer entry ("Chicken,
    thigh, raw" over "Chicken, thigh, skin removed, lean flesh, raw") is
    the more likely match for an unqualified pantry name.
    """
    query_tokens = tokenise(name)
    scored = [(food, _score(query_tokens, food)) for food in foods]
    scored = [(food, score) for food, score in scored if score >= MIN_MATCH_SCORE]
    if not scored:
        return None
    scored.sort(key=lambda pair: (-pair[1], len(pair[0].name)))
    return scored[0]


def build_foods(
    details: dict[str, str], profiles: dict[str, dict[str, float | None]]
) -> list[AfcdFood]:
    """Join the two source tables on `Public Food Key`.

    A food missing from `profiles` (or with every nutrient null — a
    handful of AFCD rows exist only as recipe components with no standalone
    values) contributes nothing and is dropped rather than kept as a
    match candidate with nothing to offer.
    """
    foods = []
    for key, name in details.items():
        nutrients = profiles.get(key)
        if not nutrients or all(v is None for v in nutrients.values()):
            continue
        foods.append(AfcdFood(key=key, name=name, tokens=tokenise(name), nutrients=nutrients))
    return foods


def _load_workbook_rows(path: Path, sheet: str):
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    rows = worksheet.iter_rows(min_row=HEADER_ROW, values_only=True)
    header = list(next(rows))
    for row in rows:
        if row[0]:  # a Public Food Key is present
            yield dict(zip(header, row))


def load_food_details(path: Path = FOOD_DETAILS_FILE) -> dict[str, str]:
    return {
        row["Public Food Key"]: row["Food Name"]
        for row in _load_workbook_rows(path, "Food details")
    }


def load_nutrient_profiles(
    path: Path = NUTRIENT_PROFILES_FILE,
) -> dict[str, dict[str, float | None]]:
    profiles = {}
    for row in _load_workbook_rows(path, "All solids & liquids per 100 g"):
        values = {}
        for field, column in NUTRIENT_COLUMNS.items():
            value = row.get(column)
            values[field] = float(value) if isinstance(value, (int, float)) else None
        values["calories_kcal"] = kcal_from_kj(values["energy_kj"])
        profiles[row["Public Food Key"]] = values
    return profiles


def load_afcd(data_dir: Path = DATA_DIR) -> list[AfcdFood]:
    """The full local dataset, or an empty list if it hasn't been
    downloaded — callers treat that as "no AFCD available", not an error,
    since the fallback (an AI estimate) still works without it."""
    food_details = data_dir / "food_details.xlsx"
    nutrient_profiles = data_dir / "nutrient_profiles.xlsx"
    if not food_details.exists() or not nutrient_profiles.exists():
        return []
    return build_foods(load_food_details(food_details), load_nutrient_profiles(nutrient_profiles))
